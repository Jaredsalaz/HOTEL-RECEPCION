from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Tuple
from io import BytesIO

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.models import Room, Reservation, ReservationStatus, RoomStatus
from app.schemas import DashboardStats

class ReportService:
    
    @staticmethod
    def get_dashboard_stats(db: Session) -> DashboardStats:
        """Get dashboard statistics"""
        # Room statistics
        total_rooms = db.query(func.count(Room.id)).scalar()
        occupied_rooms = db.query(func.count(Room.id)).filter(Room.status == RoomStatus.OCCUPIED).scalar()
        available_rooms = db.query(func.count(Room.id)).filter(Room.status == RoomStatus.AVAILABLE).scalar()
        maintenance_rooms = db.query(func.count(Room.id)).filter(Room.status == RoomStatus.MAINTENANCE).scalar()
        
        # Occupancy rate
        occupancy_rate = (occupied_rooms / total_rooms * 100) if total_rooms > 0 else 0
        
        # Reservation statistics
        active_reservations = db.query(func.count(Reservation.id)).filter(
            Reservation.status == ReservationStatus.ACTIVE
        ).scalar()
        
        # Today's check-ins/check-outs
        today = datetime.now().date()
        today_checkins = db.query(func.count(Reservation.id)).filter(
            func.date(Reservation.check_in_date) == today,
            Reservation.status.in_([ReservationStatus.PENDING, ReservationStatus.ACTIVE])
        ).scalar()
        
        today_checkouts = db.query(func.count(Reservation.id)).filter(
            func.date(Reservation.check_out_date) == today,
            Reservation.status == ReservationStatus.ACTIVE
        ).scalar()
        
        # Guest count
        from app.models import Guest
        total_guests = db.query(func.count(Guest.id)).scalar()
        
        # Revenue
        revenue_today = db.query(func.sum(Reservation.total_price)).filter(
            func.date(Reservation.created_at) == today,
            Reservation.status != ReservationStatus.CANCELLED
        ).scalar() or Decimal('0.00')
        
        # Revenue this month
        start_of_month = today.replace(day=1)
        revenue_month = db.query(func.sum(Reservation.total_price)).filter(
            Reservation.created_at >= start_of_month,
            Reservation.status != ReservationStatus.CANCELLED
        ).scalar() or Decimal('0.00')
        
        return DashboardStats(
            total_rooms=total_rooms or 0,
            occupied_rooms=occupied_rooms or 0,
            available_rooms=available_rooms or 0,
            maintenance_rooms=maintenance_rooms or 0,
            occupancy_rate=round(occupancy_rate, 2),
            active_reservations=active_reservations or 0,
            today_checkins=today_checkins or 0,
            today_checkouts=today_checkouts or 0,
            total_guests=total_guests or 0,
            revenue_today=revenue_today,
            revenue_month=revenue_month
        )
    
    @staticmethod
    def generate_occupancy_pdf(db: Session, start_date: datetime, end_date: datetime) -> BytesIO:
        """Generate occupancy report in PDF format"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        title = Paragraph("Hotel Occupancy Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))
        
        # Date range
        date_text = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Get data
        reservations = db.query(Reservation).filter(
            Reservation.check_in_date >= start_date,
            Reservation.check_in_date <= end_date
        ).all()
        
        # Summary statistics
        stats = ReportService.get_dashboard_stats(db)
        summary_data = [
            ['Metric', 'Value'],
            ['Total Rooms', str(stats.total_rooms)],
            ['Occupied Rooms', str(stats.occupied_rooms)],
            ['Available Rooms', str(stats.available_rooms)],
            ['Occupancy Rate', f'{stats.occupancy_rate}%'],
            ['Active Reservations', str(stats.active_reservations)],
        ]
        
        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.5 * inch))
        
        # Reservations detail
        elements.append(Paragraph("Reservation Details", styles['Heading2']))
        elements.append(Spacer(1, 0.2 * inch))
        
        if reservations:
            reservation_data = [['ID', 'Room', 'Guest', 'Check-in', 'Check-out', 'Status']]
            for res in reservations:
                reservation_data.append([
                    str(res.id),
                    res.room.room_number,
                    f"{res.guest.first_name} {res.guest.last_name}",
                    res.check_in_date.strftime('%Y-%m-%d'),
                    res.check_out_date.strftime('%Y-%m-%d'),
                    res.status
                ])
            
            res_table = Table(reservation_data, colWidths=[0.5*inch, 0.8*inch, 1.5*inch, 1.2*inch, 1.2*inch, 1*inch])
            res_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            elements.append(res_table)
        else:
            elements.append(Paragraph("No reservations found in this period.", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_occupancy_excel(db: Session, start_date: datetime, end_date: datetime) -> BytesIO:
        """Generate occupancy report in Excel format"""
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary['A1'] = "Hotel Occupancy Report"
        ws_summary['A1'].font = Font(size=16, bold=True, color="1a237e")
        ws_summary['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        
        # Get statistics
        stats = ReportService.get_dashboard_stats(db)
        
        # Write statistics
        row = 4
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        data = [
            ['Total Rooms', stats.total_rooms],
            ['Occupied Rooms', stats.occupied_rooms],
            ['Available Rooms', stats.available_rooms],
            ['Maintenance Rooms', stats.maintenance_rooms],
            ['Occupancy Rate', f'{stats.occupancy_rate}%'],
            ['Active Reservations', stats.active_reservations],
            ['Today Check-ins', stats.today_checkins],
            ['Today Check-outs', stats.today_checkouts],
            ['Total Guests', stats.total_guests],
            ['Revenue Today', f'${stats.revenue_today}'],
            ['Revenue This Month', f'${stats.revenue_month}'],
        ]
        
        for i, (metric, value) in enumerate(data, row + 1):
            ws_summary.cell(row=i, column=1, value=metric)
            ws_summary.cell(row=i, column=2, value=value)
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # Reservations sheet
        ws_reservations = wb.create_sheet("Reservations")
        
        # Headers
        headers = ['ID', 'Room Number', 'Guest Name', 'Email', 'Phone', 
                   'Check-in', 'Check-out', 'Guests Count', 'Total Price', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws_reservations.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Get reservations
        reservations = db.query(Reservation).filter(
            Reservation.check_in_date >= start_date,
            Reservation.check_in_date <= end_date
        ).all()
        
        # Write data
        for row, res in enumerate(reservations, 2):
            ws_reservations.cell(row=row, column=1, value=res.id)
            ws_reservations.cell(row=row, column=2, value=res.room.room_number)
            ws_reservations.cell(row=row, column=3, value=f"{res.guest.first_name} {res.guest.last_name}")
            ws_reservations.cell(row=row, column=4, value=res.guest.email)
            ws_reservations.cell(row=row, column=5, value=res.guest.phone)
            ws_reservations.cell(row=row, column=6, value=res.check_in_date.strftime('%Y-%m-%d'))
            ws_reservations.cell(row=row, column=7, value=res.check_out_date.strftime('%Y-%m-%d'))
            ws_reservations.cell(row=row, column=8, value=res.guests_count)
            ws_reservations.cell(row=row, column=9, value=float(res.total_price))
            ws_reservations.cell(row=row, column=10, value=res.status)
        
        # Adjust column widths
        for col in range(1, 11):
            ws_reservations.column_dimensions[chr(64 + col)].width = 15
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def get_room_status_report(db: Session) -> Dict:
        """Get current status of all rooms"""
        rooms = db.query(Room).all()
        
        report = {
            'timestamp': datetime.now().isoformat(),
            'total_rooms': len(rooms),
            'rooms_by_status': {},
            'rooms_by_type': {},
            'rooms_detail': []
        }
        
        # Group by status
        for status in RoomStatus:
            count = len([r for r in rooms if r.status == status])
            report['rooms_by_status'][status.value] = count
        
        # Group by type
        from app.models import RoomType
        for room_type in RoomType:
            count = len([r for r in rooms if r.type == room_type])
            report['rooms_by_type'][room_type.value] = count
        
        # Detail
        for room in rooms:
            report['rooms_detail'].append({
                'id': room.id,
                'room_number': room.room_number,
                'type': room.type,
                'status': room.status,
                'price_per_night': float(room.price_per_night),
                'capacity': room.capacity,
                'floor': room.floor
            })
        
        return report
